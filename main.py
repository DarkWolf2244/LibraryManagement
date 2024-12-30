import tkinter as tk
from tkinter import messagebox
from typing import List
import openpyxl
from PIL import Image, ImageTk

# Colors and dimensions
COLORS = {
    "surface": "#F2E5BF",
    "surface_2": "#CB6040",
    "primary": "#257180",
    "danger": "black"
}

WINDOW_DIMENSIONS = {
    "login": (600, 400),
    "home": (1000, 700),
    "create_book": (500, 500),
    "create_customer": (500, 500),
    "view_book": (500, 500),
    "view_customer": (500, 500),
    "issue_book": (400, 600)
}

# User credentials and login bypass
VALID_CREDENTIALS = {"Admin": "123"}
BYPASS_LOGIN = False
DATABASE_PATH = "database.xlsx"
LOGGED_IN_USER = (None, None)
# Globals
main_frame = None
data = None
workbook = None
issue_window = None

class Customer:
    def __init__(self, name, book_borrowed="NULL"):
        self.name = name
        self.book_borrowed = book_borrowed

        
    def serialize(self):
        return (self.name, self.book_borrowed)
    
class Book:
    def __init__(self, title, author, isbn, issued_to=""):
        self.title = title
        self.author = author
        self.isbn = isbn
        self.issued_to = issued_to

    def serialize(self):
        return (self.isbn, self.title, self.author, self.issued_to)

class Data:
    _customers: List[Customer] = []
    _books: List[Book] = []

    def __init__(self, database):
        self.database = database

    @property
    def customers(self):
        return self._customers

    @customers.setter
    def customers(self, value):
        self._customers = value
        self.database.save_to_file()

    @property
    def books(self):
        return self._books

    @books.setter
    def books(self, value):
        self._books = value
        self.database.save_to_file()

class Database:
    def __init__(self, file_path: str):
        self.data = Data(self)
        self.workbook = None
        self.file_path = file_path

    def load_from_file(self):
        self.workbook = openpyxl.load_workbook(self.file_path)
        books_sheet = self.workbook['Books']
        customers_sheet = self.workbook["Customers"]

        books = []
        for isbn, title, author, checked_out_by in books_sheet.iter_rows(min_row=2, values_only=True):
            books.append(Book(title=title, author=author, isbn=isbn, issued_to=checked_out_by))
        
        customers = []
        for name, book_borrowed in customers_sheet.iter_rows(min_row=2, values_only=True):
            customers.append(Customer(name=name, book_borrowed=book_borrowed))

        self.workbook.close()
        self.data.books = books
        self.data.customers = customers
    
    def save_to_file(self):
        self.workbook = openpyxl.load_workbook(self.file_path)
        books_sheet = self.workbook['Books']
        customers_sheet = self.workbook["Customers"]

        for row, book in enumerate(self.data.books, start=2):
            for col, value in enumerate(book.serialize(), start=1):
                books_sheet.cell(row=row, column=col, value=value)
        
        for row, customer in enumerate(self.data.customers, start=2):
            for col, value in enumerate(customer.serialize(), start=1):
                customers_sheet.cell(row=row, column=col, value=value)

        self.workbook.save(self.file_path)
        self.workbook.close()

class VerticalScrolledFrame(tk.Frame):
    """Scrollable frame with vertical scrolling."""
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self.canvas = tk.Canvas(self, bg=COLORS["surface"], bd=0, highlightthickness=0)
        self.scrollbar = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.interior = tk.Frame(self.canvas, bg=COLORS["surface"], padx=25, pady=25)

        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.create_window((0, 0), window=self.interior, anchor="nw")

        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.interior.bind("<Configure>", self._update_scroll_region)
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)

    def _update_scroll_region(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(-int(event.delta / 120), "units")

def load_data():
    global data
    global workbook

    """Load data from the Excel file."""
    workbook = openpyxl.load_workbook(DATABASE_PATH)
    books_sheet = workbook['Books']
    customers_sheet = workbook["Customers"]

    books = []
    for row in books_sheet.iter_rows(min_row=2, values_only=True):
        books.append(row)
    
    customers = []
    for row in customers_sheet.iter_rows(min_row=2, values_only=True):
        customers.append(row)

    workbook.close()
    data = {
        'books': books,
        'customers': customers
    }

db = Database(DATABASE_PATH)
db.load_from_file()

def center_window(window, width, height):
    """Center a window on the screen."""
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f"{width}x{height}+{x}+{y}")
    
def create_root():
    """Create and configure the main application window."""
    root = tk.Tk()
    root.title("Login | Library Management System")
    root.configure(bg=COLORS["surface"], padx=10, pady=10)
    center_window(root, *WINDOW_DIMENSIONS["login"])
    return root

def display_login(root):
    global LOGGED_IN_USER

    """Display the login page."""
    def handle_login():
        global LOGGED_IN_USER

        username, password = username_entry.get(), password_entry.get()
        if BYPASS_LOGIN or VALID_CREDENTIALS.get(username) == password:
            LOGGED_IN_USER = (username, password)
            root.destroy()
            open_home_page()
        else:
            messagebox.showerror("Login Failed", "Invalid username or password.")

    tk.Label(root, text="Library Management System", font=("Georgia", 24), 
             bg=COLORS["surface"], fg=COLORS["primary"]).pack(pady=20)
    

    for label_text, show_char in [("Username", None), ("Password", "*")]:
        tk.Label(root, text=label_text, font=("Georgia", 12), bg=COLORS["surface"]).pack(pady=10)
        entry = tk.Entry(root, show=show_char)
        entry.pack(pady=10)
        if label_text == "Username":
            username_entry = entry
        else:
            password_entry = entry

    tk.Button(root, text="Login", font=("Georgia", 12), bg=COLORS["primary"], 
              fg="white", command=handle_login).pack(pady=20)

def open_home_page():
    """Open the home page."""
    home = tk.Tk()
    home.title("Home | Library Management System")
    home.configure(bg=COLORS["surface"])
    center_window(home, *WINDOW_DIMENSIONS["home"])
    display_home_page(home, "default")
    home.mainloop()

def open_create_customer_page():
    window = tk.Toplevel()
    window.title("Add Customer | Library Management System")
    window.configure(bg=COLORS["surface"])
    center_window(window, *WINDOW_DIMENSIONS["create_customer"])
    window.focus_force()
    tk.Label(window, text="Add a Customer", font=("Georgia", 24), bg=COLORS["surface_2"], fg=COLORS["surface"]).pack(fill="x",)

    entries = {}
    for label_text in ["Customer Name"]:
        tk.Label(window, text=label_text, font=("Georgia", 12), bg=COLORS["surface"]).pack(pady=10)
        entry = tk.Entry(window, font=("Georgia", 12))
        entry.pack(pady=10)
        entries[label_text.lower()] = entry

    def create_customer():
        if all(entry.get().strip() for entry in entries.values()):
            messagebox.showinfo("Success", "Customer added successfully! Refresh the page to view.")
            db.data.customers.append(entries['customer name'].get())
            db.save_to_file()
            window.destroy()
            load_data()

        else:
            messagebox.showerror("Error", "All fields must be filled.")
            window.focus_force()

    tk.Button(window, text="Add Customer", font=("Georgia", 12), 
              bg=COLORS["primary"], fg="white", command=create_customer).pack(pady=20)
    
def display_home_page(home, page):
    """Display the home page with a sidebar and dynamic content."""
    global main_frame

    def switch_page(new_page):
        global main_frame
        main_frame.destroy()
        main_frame = PAGE_GENERATORS.get(new_page, generate_default_frame)(home)
        main_frame.pack(side="right", fill="both", expand=True)

    sidebar = tk.Frame(home, bg=COLORS["surface_2"], width=200, padx=5)
    sidebar.pack(side="left", fill="y")

    buttons_frame = tk.Frame(sidebar, bg=COLORS["surface_2"])
    buttons_frame.pack(fill="y", expand=True, pady=50)

    logout_frame = tk.Frame(sidebar, bg=COLORS["surface_2"])
    logout_frame.pack(fill="x")

    buttons = [
        ("List Books", lambda: switch_page("books"), "primary"),
        ("List Customers", lambda: switch_page("customers"), "primary"),
        ("Add Book", open_create_book_page, "primary"),
        ("Add Customer", open_create_customer_page, "primary"),
    ]

    for text, command, color in buttons:
        tk.Button(buttons_frame, text=text, font=("Georgia", 12), bg=COLORS[color], 
                  fg="white", command=command).pack(fill="x", pady=10)

    tk.Button(logout_frame, text="Logout", font=("Georgia", 12), bg=COLORS["primary"], 
              fg="white", command=lambda: home.destroy()).pack(fill="x", pady=10)
    
    main_frame = PAGE_GENERATORS.get(page, generate_default_frame)(home)
    main_frame.pack(side="right", fill="both", expand=True)

def generate_default_frame(parent):
    """Generate the default frame content."""
    frame = tk.Frame(parent, bg=COLORS["surface"])
    
    splash_img = Image.open("splash.png").resize((1000, 800), Image.Resampling.LANCZOS)
    photo = ImageTk.PhotoImage(splash_img)
    label = tk.Label(frame, image=photo)
    label.place(x=0, y=0, relwidth=1, relheight=1)
    label.image = photo

    tk.Label(frame, text=f"Welcome, {LOGGED_IN_USER[0]}.", font=("Georgia", 24), 
             bg=COLORS["surface_2"], fg=COLORS["surface"]).pack( fill="x")
    
    tk.Label(frame, text="Library Management Software", font=("Georgia", 24), 
             fg=COLORS["primary"], bg=COLORS["surface"]).place(relx=0.5, rely=0.3, anchor="center")
   

    return frame

def issue_book(book_customer_frame):
    book, customer, frame = book_customer_frame
    customer: Customer = customer

    customer_index = db.data.customers.index(customer)
    book_index = db.data.books.index(book)

    if customer.book_borrowed != "NULL":
        borrowed_book = next((b for b in db.data.books if b.isbn == customer.book_borrowed), None)

        user_input = messagebox.askyesno("Info", f"Customer {customer.name} has already borrowed a book, named '{borrowed_book.title}'. Do you want to return the book and then issue '{borrowed_book.title}' to {customer.name}?")
        if not user_input:
            return                

        db.data.books[db.data.books.index(borrowed_book)].issued_to = "NULL"
        db.save_to_file()

    
    db.data.customers[customer_index].book_borrowed = book.isbn
    db.data.books[book_index].issued_to = customer.name
    db.save_to_file() # You can also just set db.books to something, that automatically saves, but we do it manually here because mutating a list member does not count as reassigning the list



    messagebox.showinfo(f"Issued '{book.title}'", f"Successfully issued '{book.title}' to {customer.name}.")
    frame.destroy()
    issue_window.destroy()



def issue_search(entry, frame, book, window):


    for widget in frame.winfo_children():
        widget.destroy()

    found_customers = False
    for customer in db.data.customers:
        if customer.name.lower().startswith(entry.get().lower()):
            found_customers = True
            tk.Button(frame, text=customer.name, bg=COLORS["primary"], fg=COLORS['surface'], font=("Georgia", 12), command= lambda book_customer=(book, customer, window): issue_book(book_customer)).pack(padx=20, pady=20)

    if not found_customers:
        tk.Label(frame, text="No customers found with that name.", bg=COLORS["surface"], fg=COLORS['surface_2'], font=("Georgia", 12)).pack(padx=20, pady=20)

def open_issue_window(book: Book, main_window):
    global issue_window
    window = tk.Toplevel()
    window.title(f"{book.title} | Issue")
    window.configure(bg=COLORS["surface"])
    center_window(window, *WINDOW_DIMENSIONS["issue_book"])
    window.focus_force()

    issue_window = window

        
    tk.Label(window, text="Issue to a customer...", bg=COLORS["surface"], fg=COLORS['primary'], font=("Georgia", 24)).pack(padx=20, pady=20)

    search_bar = tk.Frame(window, bg=COLORS["surface"])
    search_bar.pack(fill="x", pady=5, padx=10)

    # Add the text entry box
    search_entry = tk.Entry(search_bar, font=("Georgia", 12))
    search_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))

    customer_frame = VerticalScrolledFrame(window, bg=COLORS['surface'])
    customer_frame.pack(fill="y", expand=True)

    search_button = tk.Button(search_bar, text="Search", font=("Georgia", 12), 
                               bg=COLORS["primary"], fg="white", command=lambda entry=search_entry: issue_search(entry, customer_frame.interior, book, main_window))
    search_button.pack(side="right")

    for customer in db.data.customers:
        tk.Button(customer_frame.interior, text=customer.name, bg=COLORS["primary"], fg=COLORS['surface'], font=("Georgia", 12), command= lambda book_customer=(book, customer, main_window): issue_book(book_customer)).pack(padx=20, pady=20, fill="x", anchor="center")
        

def open_delete_book_prompt(book, window):
    if book.issued_to != "NULL":
        messagebox.showerror("Book currently issued", "Sorry, you can't delete a book that is issued. Return the book and try again.")
        return
    
    user_input = messagebox.askyesno("Delete Book", "Are you sure you want to delete this book?")
    if user_input:
        db.data.books.remove(book)
        db.save_to_file()
        messagebox.showinfo("Deleted Book", "Successfully deleted book. Refresh page to view.")
        load_data()
        window.destroy()
        return

def open_return_book_prompt(book, window):
    user_input = messagebox.askyesno("Return Book", "Are you sure you want to return this book?")
    if user_input:
        customer_index = db.data.customers.index(next((c for c in db.data.customers if c.name == book.issued_to), None))

        db.data.books[db.data.books.index(book)].issued_to = "NULL"

        db.data.customers[customer_index].book_borrowed = "NULL"
        db.save_to_file()

        messagebox.showinfo("Returned Book", "Successfully returned book.")
        window.destroy()
        return
        

def open_book_page(book: Book):
    window = tk.Toplevel()
    window.title(f"{book.title} | Details")
    window.configure(bg=COLORS["surface"])
    center_window(window, *WINDOW_DIMENSIONS["view_book"])
    window.focus_force()

    tk.Label(window, text=f"{book.title}", font=("Georgia", 24), bg=COLORS["surface_2"], fg=COLORS["surface"]).pack(fill="x",)
    tk.Label(window, text="Book Details", font=("Georgia", 12), bg=COLORS["surface"], fg=COLORS["primary"]).pack(pady=30)
    
    tk.Label(window, text=f"ISBN: {book.isbn}", font=("Georgia", 12), bg=COLORS["surface"], fg=COLORS["primary"]).pack(anchor="w", padx=10, pady=10)
    tk.Label(window, text=f"Title: {book.title}", font=("Georgia", 12), bg=COLORS["surface"], fg=COLORS["primary"]).pack(anchor="w", padx=10, pady=10)
    tk.Label(window, text=f"Author: {book.author}", font=("Georgia", 12), bg=COLORS["surface"], fg=COLORS["primary"]).pack(anchor="w", padx=10, pady=10)
    
    tk.Label(window, text=f"This book is currently {'available' if book.issued_to == 'NULL' else f'borrowed by {book.issued_to}'}.", font=("Georgia", 12), bg=COLORS["surface"], fg=COLORS["primary"]).pack(pady=30)
    if book.issued_to != "NULL":
        tk.Button(window, text="Open Customer", bg=COLORS["primary"], fg=COLORS["surface"], font=("Georgia", 12), command=lambda book=book: open_customer_page(next((c for c in db.data.customers if c.name == book.issued_to), None))).pack(pady=10)
    
    tk.Button(window, text="Issue Book", bg=COLORS["primary"], fg=COLORS["surface"], font=("Georgia", 12), command=lambda book=book: open_issue_window(book, window)).pack(pady=10)
    
    if book.issued_to != "NULL":
        tk.Button(window, text="Return Book", bg=COLORS["primary"], fg=COLORS["surface"], font=("Georgia", 12), command=lambda book=book: open_return_book_prompt(book, window)).pack(pady=10)
    tk.Button(window, text="Delete Book", bg=COLORS["danger"], fg=COLORS["surface"], font=("Georgia", 12), command=lambda book=book: open_delete_book_prompt(book, window)).pack(pady=10)

def generate_list_books_frame(parent):
    """Generate the frame displaying a grid of books."""
    frame = tk.Frame(parent, bg=COLORS["surface"])
    
    tk.Label(frame, text="Book Catalogue", font=("Georgia", 24), bg=COLORS["surface_2"], fg=COLORS["surface"]).pack(fill="x",)

    search_bar = tk.Frame(frame, bg=COLORS["surface"])
    search_bar.pack(fill="x", pady=5, padx=10)

    # Add the text entry box
    search_entry = tk.Entry(search_bar, font=("Georgia", 12))
    search_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

    # Add the search button
    search_button = tk.Button(search_bar, text="Search", font=("Georgia", 12), 
                               bg=COLORS["primary"], fg="white", command= lambda: messagebox.showerror("Not yet implemented.", "Search functionality has not been implemented yet."))
    search_button.pack(side="right")

    scrollable = VerticalScrolledFrame(frame)
    scrollable.pack(fill="both", expand=True)
    

    for i, book in enumerate(db.data.books):
        col = i % 3
        row = i // 3

        bg_color = COLORS["primary"] if i % 2 == 0 else COLORS["surface_2"]
        tk.Button(scrollable.interior, text=f"{book.title}", 
                      font=("Georgia", 12), bg=bg_color, 
                      fg="white", width=25, command=lambda book=book: open_book_page(book) ).grid(row=row, column=col, padx=5, pady=5)
    

    return frame

def open_delete_customer_prompt(customer, window):
    if messagebox.askyesno("Delete Customer", "Are you sure you want to delete this customer?"):
        if customer.book_borrowed != "NULL":
            messagebox.showerror("Book currently borrowed", "Sorry, you can't delete a customer that has borrowed a book. Return the book and try again.")
            return

        db.data.customers.remove(customer)
        db.save_to_file()

        messagebox.showinfo("Deleted Customer", "Successfully deleted customer. Refresh page to view.")
        load_data()
        window.destroy()
        return

def open_customer_page(customer: Customer):
    window = tk.Toplevel()
    window.title(f"{customer.name} | Details")
    window.configure(bg=COLORS["surface"])
    center_window(window, *WINDOW_DIMENSIONS["view_customer"])
    window.focus_force()

    book_borrowed = None
    label1_text = "This customer has not borrowed any book."

    if customer.book_borrowed != "NULL":
        book_borrowed = next((b for b in db.data.books if b.isbn == customer.book_borrowed), None)
        label1_text = f"Book borrowed: {book_borrowed.title}."
    

    tk.Label(window, text=f"{customer.name}", font=("Georgia", 24), bg=COLORS["surface_2"], fg=COLORS["surface"]).pack(fill="x",)
    tk.Label(window, text="Customer Details", font=("Georgia", 12), bg=COLORS["surface"], fg=COLORS["primary"]).pack(pady=30)
    
    tk.Label(window, text=f"Name: {customer.name}", font=("Georgia", 12), bg=COLORS["surface"], fg=COLORS["primary"]).pack(anchor="w", padx=10, pady=10)
    tk.Label(window, text=label1_text, font=("Georgia", 12), bg=COLORS["surface"], fg=COLORS["primary"]).pack(anchor="w", padx=10, pady=10)
    if book_borrowed:
        tk.Button(window, text="Open Book", bg=COLORS["primary"], fg=COLORS["surface"], font=("Georgia", 12), command=lambda book=book_borrowed: open_book_page(book)).pack(pady=10)
    tk.Button(window, text="Delete Customer", bg=COLORS["danger"], fg=COLORS["surface"], font=("Georgia", 12), command=lambda customer=customer: open_delete_customer_prompt(customer, window)).pack(pady=10)

def generate_list_customers_frame(parent):
    """Generate a placeholder frame for customer management."""
    frame = tk.Frame(parent, bg=COLORS["surface"])
    
    tk.Label(frame, text="Customer List", font=("Georgia", 24), bg=COLORS["surface_2"], fg=COLORS["surface"]).pack(fill="x",)

    search_bar = tk.Frame(frame, bg=COLORS["surface"])
    search_bar.pack(fill="x", pady=5, padx=10)

    # Add the text entry box
    search_entry = tk.Entry(search_bar, font=("Georgia", 12))
    search_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

    # Add the search button
    search_button = tk.Button(search_bar, text="Search", font=("Georgia", 12), 
                               bg=COLORS["primary"], fg="white", command= lambda: messagebox.showerror("Not yet implemented.", "Search functionality has not been implemented yet."))
    search_button.pack(side="right")

    scrollable = VerticalScrolledFrame(frame)
    scrollable.pack(fill="both", expand=True)
    

    for i, customer in enumerate(db.data.customers):
        col = i % 3
        row = i // 3

        bg_color = COLORS["primary"] if i % 2 == 0 else COLORS["surface_2"]
        tk.Button(scrollable.interior, text=f"{customer.name}", 
                      font=("Georgia", 12), bg=bg_color, 
                      fg="white", width=25, command=lambda customer=customer: open_customer_page(customer)).grid(row=row, column=col, padx=5, pady=5)
    
    return frame

def open_create_book_page():
    """Open the 'Create Book' window."""
    window = tk.Toplevel()
    window.title("Add a Book")
    window.configure(bg=COLORS["surface"])
    center_window(window, *WINDOW_DIMENSIONS["create_book"])
    window.focus_force()

    tk.Label(window, text="Create Book", font=("Georgia", 24), bg=COLORS["surface_2"], fg=COLORS["surface"]).pack(fill="x",)
    entries = {}
    for label_text in ["Title", "Author", "ISBN"]:
        tk.Label(window, text=label_text, font=("Georgia", 12), bg=COLORS["surface"]).pack(pady=10)
        entry = tk.Entry(window, font=("Georgia", 12))
        entry.pack(pady=10)
        entries[label_text.lower()] = entry

    def create_book():
        if all(entry.get().strip() for entry in entries.values()):
            messagebox.showinfo("Success", "Book created successfully! Refresh the page to view.")
            db.data.books.append(Book(entries['title'].get(), entries['author'].get(), entries['isbn'].get()))
            db.save_to_file()

            window.destroy()
        else:
            messagebox.showerror("Error", "All fields must be filled.")
            window.focus_force()
    tk.Button(window, text="Create Book", font=("Georgia", 12), 
              bg=COLORS["primary"], fg="white", command=create_book).pack(pady=20)

    
PAGE_GENERATORS = {
    "default": generate_default_frame,
    "books": generate_list_books_frame,
    "customers": generate_list_customers_frame
}

def run_app():
    """Run the application."""
    load_data()

    if not BYPASS_LOGIN:
        root = create_root()
        display_login(root)
        root.mainloop()
    else:
        open_home_page()

if __name__ == "__main__":
    run_app()
